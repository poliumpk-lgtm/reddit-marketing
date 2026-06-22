"""Reddit lead-gen orchestrator (template).

Three commands:
  scan           — pull RSS for the configured subreddit list, score each post
                   against scoring.yml + entities.yml, bucket into
                   Релевантно/Околорелевантно/Нерелевантно, enrich the
                   non-irrelevant ones via SB-CDP, write xlsx with 3 sheets.

  sync-overrides — re-read the latest xlsx, apply Manual Override column,
                   move re-classified posts between sheets, append
                   classification_feedback to lessons.jsonl.

  sync-voice     — re-read the latest xlsx, diff Suggested Reply ↔ Final Reply
                   for each row, append voice_feedback to lessons.jsonl,
                   refresh voice/examples.jsonl with the latest 10 actual
                   replies for next-scan few-shot.

The "scan" command does NOT generate Suggested Reply text — it only writes
the bucket, score, score_breakdown, and a stub like "DRAFT: <use_case>"
into Suggested Reply. Real reply text is filled in afterwards (either by
an LLM call or a human in fill_replies.py).

Files alongside this script:
  scoring.yml       — rubric (intent, subreddit_context, anti_fits, etc.)
  entities.yml      — competitors / antidetect_browsers / vendor_blogs / domain_keywords
  voice/rules.md    — voice digest (read by future draft generator)
  voice/lessons.jsonl — append-only feedback log
  voice/examples.jsonl — last N (Final Reply ≠ Suggested Reply) for few-shot

Subreddits to scan are hard-coded in this file (SUBS list). Edit there
to add/remove subs — they pair with `subreddit_buckets` in scoring.yml.
"""

import argparse
import dataclasses
import datetime as dt
import difflib
import io
import json
import re
import subprocess
import sys
from pathlib import Path
from typing import Any

import yaml
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

HERE = Path(__file__).resolve().parent
VOICE = HERE / "voice"
RSS_SCRIPT = HERE.parent / "reddit_rss.py"
ENRICH_SCRIPT = HERE.parent / "reddit_enrich.py"

SCORING_PATH = HERE / "scoring.yml"
ENTITIES_PATH = HERE / "entities.yml"
LESSONS_PATH = VOICE / "lessons.jsonl"
EXAMPLES_PATH = VOICE / "examples.jsonl"
RULES_PATH = VOICE / "rules.md"

# Subreddit list with pages-per-pull. Tune `pages` so that 100*pages posts
# comfortably covers your scan window (e.g. 72h). Sub names must match the
# `subreddit_buckets` keys (lowercased) in scoring.yml.
SUBS: list[tuple[str, int]] = [
    # ("YourTargetSubreddit", 1),
    # ("AnotherSub", 2),
]


# =============================================================================
# Config loading
# =============================================================================

def load_yaml(path: Path) -> dict:
    return yaml.safe_load(path.read_text(encoding="utf-8"))


def load_lessons() -> list[dict]:
    if not LESSONS_PATH.exists():
        return []
    return [json.loads(line) for line in LESSONS_PATH.read_text(encoding="utf-8").splitlines() if line.strip()]


def append_lesson(record: dict) -> None:
    LESSONS_PATH.parent.mkdir(parents=True, exist_ok=True)
    with LESSONS_PATH.open("a", encoding="utf-8") as f:
        f.write(json.dumps(record, ensure_ascii=False) + "\n")


# =============================================================================
# Scoring engine
# =============================================================================

@dataclasses.dataclass
class ScoreResult:
    score: float
    breakdown: list[tuple[str, float]]   # [(signal_label, weight), ...]
    bucket: str                           # relevant | borderline | irrelevant
    matched_entities: list[str]
    use_cases: list[str]
    anti_fits: list[str]                  # which anti-fits triggered (force irrelevant)


class Scorer:
    def __init__(self, scoring: dict, entities: dict, vendor_authors: set[str]):
        self.scoring = scoring
        self.entities = entities
        self.vendor_authors = vendor_authors

        # Compile patterns once.
        self._intent_patterns = {
            name: [re.compile(p, re.IGNORECASE | re.MULTILINE) for p in cfg.get("patterns", [])]
            for name, cfg in scoring.get("intent_signals", {}).items()
        }
        self._anti_fit_patterns = {
            name: [re.compile(p, re.IGNORECASE | re.MULTILINE) for p in cfg.get("patterns", [])]
            for name, cfg in scoring.get("anti_fits", {}).items()
            if "patterns" in cfg
        }
        self._negative_patterns = {
            name: [re.compile(p, re.IGNORECASE | re.MULTILINE) for p in cfg.get("patterns", [])]
            for name, cfg in scoring.get("negative_signals", {}).items()
            if "patterns" in cfg
        }

        # Subreddit → context bucket (proxy_core / competitor_brand / etc.)
        self._sub_context: dict[str, str] = {}
        for ctx, subs in scoring.get("subreddit_buckets", {}).items():
            for slug in subs:
                self._sub_context[slug.lower()] = ctx

        # Entity matchers
        def _aliases(group):
            out = []
            for ent in group:
                if isinstance(ent, dict):
                    name = ent["name"]
                    for alias in ent.get("aliases", [name]):
                        out.append((name, alias.lower()))
                else:
                    out.append((ent, ent.lower()))
            return out

        self._competitors = _aliases(entities.get("competitors", []) or [])
        self._antidetect = _aliases(entities.get("antidetect_browsers", []) or [])
        self._domain_keywords = [k.lower() for k in entities.get("domain_keywords", []) or []]

    def _match_entities(self, text: str) -> tuple[list[str], list[str]]:
        """Return (competitors_mentioned, antidetect_mentioned)."""
        tl = text.lower()
        comps = sorted({name for name, alias in self._competitors if _wb_contains(tl, alias)})
        anti = sorted({name for name, alias in self._antidetect if _wb_contains(tl, alias)})
        return comps, anti

    def _detect_use_cases(self, text: str, antidetect_mentioned: list[str]) -> list[str]:
        cases = set()
        if antidetect_mentioned:
            cases.add("multi-account")
        tl = text.lower()
        if re.search(r"\bscrap(ing|er|ed|e)\b", tl):
            cases.add("scraping")
        if re.search(r"\b(sneaker|drop|raffle|cop|copping|aco|aio)\b", tl):
            cases.add("sneaker")
        if re.search(r"\b(affiliate|arbitrage|арбитраж)\b", tl):
            cases.add("affiliate")
        if re.search(r"\b(seo|search engine optimi[sz]ation|backlink)\b", tl):
            cases.add("seo")
        if re.search(r"\b(price\s+monitor|price\s+tracking)\b", tl):
            cases.add("ecommerce-price")
        if re.search(r"\b(ad\s+verif|ad\s+test)\b", tl):
            cases.add("ad-verification")
        return sorted(cases)

    def score_post(self, post: dict) -> ScoreResult:
        title = post.get("title") or ""
        body = post.get("body_text") or ""
        text = f"{title}\n{body}"
        author = (post.get("author") or "").lstrip("/u/")
        sub = (post.get("subreddit") or "").lower()

        breakdown: list[tuple[str, float]] = []
        anti_fit_triggered: list[str] = []

        # 1. Anti-fits (any → bucket=irrelevant later)
        if author in self.vendor_authors:
            anti_fit_triggered.append("vendor_self_promo")

        for name, patterns in self._anti_fit_patterns.items():
            if any(p.search(text) for p in patterns):
                anti_fit_triggered.append(name)

        # 2. Off-topic-landed: post in proxy_core sub but NO domain_keywords
        if self._sub_context.get(sub) == "proxy_core":
            if not any(kw in text.lower() for kw in self._domain_keywords):
                anti_fit_triggered.append("off_topic_landed")

        # 3. Intent signals
        for name, patterns in self._intent_patterns.items():
            if any(p.search(text) for p in patterns):
                weight = self.scoring["intent_signals"][name]["weight"]
                breakdown.append((name, weight))

        # 4. Computed: named_competitor + use_case_match
        competitors, antidetect = self._match_entities(text)
        if competitors:
            w = self.scoring["computed_signals"]["named_competitor"]["weight"]
            breakdown.append((f"named_competitor: {', '.join(competitors)}", w))

        use_cases = self._detect_use_cases(text, antidetect)
        if use_cases:
            w = self.scoring["computed_signals"]["use_case_match"]["weight"]
            label = f"use_case_match: {', '.join(use_cases)}"
            if antidetect:
                label += f" (antidetect: {', '.join(antidetect)})"
            breakdown.append((label, w))

        # 5. Subreddit context bonus
        ctx = self._sub_context.get(sub)
        if ctx:
            w = self.scoring["subreddit_context"][ctx]
            breakdown.append((f"sub_context: {ctx}", w))

        # 6. Negative pattern signals
        for name, patterns in self._negative_patterns.items():
            if any(p.search(text) for p in patterns):
                w = self.scoring["negative_signals"][name]["weight"]
                breakdown.append((name, w))

        # Bucketing
        if anti_fit_triggered:
            score = 0.0
            bucket = "irrelevant"
        else:
            score = sum(w for _, w in breakdown)
            cuts = self.scoring["bucket_cutoffs"]
            if score >= cuts["relevant"]:
                bucket = "relevant"
            elif score >= cuts["borderline"]:
                bucket = "borderline"
            else:
                bucket = "irrelevant"

        return ScoreResult(
            score=score,
            breakdown=breakdown,
            bucket=bucket,
            matched_entities=competitors + antidetect,
            use_cases=use_cases,
            anti_fits=anti_fit_triggered,
        )


def _wb_contains(text: str, alias: str) -> bool:
    """Substring match with word boundaries when feasible."""
    if not alias:
        return False
    # If alias has special chars or short — fall back to bare substring
    if len(alias) <= 3 or not all(c.isalnum() or c in "- " for c in alias):
        return alias in text
    pat = r"\b" + re.escape(alias) + r"\b"
    return bool(re.search(pat, text))


def format_breakdown(score: float, br: list[tuple[str, float]], anti_fits: list[str]) -> str:
    if anti_fits:
        return f"ANTI-FIT: {', '.join(anti_fits)} → score=0 (irrelevant)"
    parts = [f"{w:+g} ({label})" for label, w in br]
    return " ".join(parts) + f" = {score:g}"


# =============================================================================
# Cross-poster detection
# =============================================================================

def mark_cross_posts(posts: list[dict], scorer: Scorer) -> None:
    """For (author, normalized_title) groups with >1 post within 24h,
    mark all but the most-engaged one with cross_post_duplicate penalty."""
    groups: dict[tuple[str, str], list[dict]] = {}
    for p in posts:
        author = (p.get("author") or "").lstrip("/u/").lower()
        title_norm = re.sub(r"\W+", " ", (p.get("title") or "").lower()).strip()
        if not author or not title_norm:
            continue
        key = (author, title_norm)
        groups.setdefault(key, []).append(p)

    weight = scorer.scoring["negative_signals"]["cross_post_duplicate"]["weight"]
    for key, group in groups.items():
        if len(group) < 2:
            continue
        # Pick the post with max (upvotes + comments) — the "winner" keeps full score.
        def engagement(p):
            return (p.get("upvotes") or 0) + (p.get("comments") or 0)

        winner = max(group, key=engagement)
        for p in group:
            if p is winner:
                continue
            sr: ScoreResult = p["_score"]
            if "cross_post_duplicate" in {label.split(":")[0] for label, _ in sr.breakdown}:
                continue
            sr.breakdown.append(("cross_post_duplicate", weight))
            sr.score += weight
            # Re-bucket
            if sr.score < 2:
                sr.bucket = "irrelevant"
            elif sr.score < 6:
                sr.bucket = "borderline"


# =============================================================================
# RSS pull
# =============================================================================

def pull_sub(slug: str, pages: int) -> list[dict]:
    cmd = [
        sys.executable, str(RSS_SCRIPT),
        "--subreddit", slug, "--feed", "new",
        "--limit", "100", "--pages", str(pages),
    ]
    proc = subprocess.run(cmd, capture_output=True, text=True, timeout=180)
    if proc.returncode != 0:
        print(f"  ! {slug}: rc={proc.returncode} stderr={proc.stderr.strip()[:200]}",
              file=sys.stderr)
        return []
    try:
        return json.loads(proc.stdout)
    except json.JSONDecodeError as e:
        print(f"  ! {slug}: bad JSON ({e})", file=sys.stderr)
        return []


def filter_recent(posts: list[dict], hours: int) -> list[dict]:
    cutoff = dt.datetime.now(dt.timezone.utc) - dt.timedelta(hours=hours)
    out = []
    for p in posts:
        try:
            when = dt.datetime.fromisoformat(p.get("published") or "")
        except ValueError:
            continue
        if when >= cutoff:
            out.append(p)
    return out


# =============================================================================
# Enrichment
# =============================================================================

def enrich_posts(posts: list[dict]) -> dict[str, dict]:
    """Run reddit_enrich.py on the given posts. Returns {url: enrich_data}."""
    if not posts:
        return {}
    payload = json.dumps([{"url": p["url"]} for p in posts], ensure_ascii=False)
    cmd = [sys.executable, str(ENRICH_SCRIPT)]
    proc = subprocess.run(cmd, input=payload, capture_output=True, text=True, timeout=900)
    if proc.returncode != 0:
        print(f"  ! enrich: rc={proc.returncode}", file=sys.stderr)
        print(proc.stderr[-2000:], file=sys.stderr)
        return {}
    # Stdout may have SB framework noise mixed in. Find the JSON block.
    m = re.search(r"\n\[\s*\n", proc.stdout)
    if m:
        json_blob = proc.stdout[m.start():]
    else:
        json_blob = proc.stdout
    try:
        enriched = json.loads(json_blob)
    except json.JSONDecodeError:
        print(f"  ! enrich: failed to parse stdout", file=sys.stderr)
        return {}
    return {e["url"]: e for e in enriched}


# =============================================================================
# Workbook I/O
# =============================================================================

DRAFT_COLS = [
    "Published At", "Subreddit", "Title", "URL", "Summary",
    "Suggested Reply", "Final Reply", "Status", "Score", "Score Breakdown",
    "Bucket", "Manual Override", "Notes", "My Hypothesis",
    "Author", "Author Profile URL", "Upvotes", "Comments",
]

REJECTED_COLS = DRAFT_COLS  # same schema; just empty Suggested Reply / Final Reply / Score Breakdown


def fmt_published(p: dict) -> str:
    try:
        when = dt.datetime.fromisoformat(p["published"])
        return when.strftime("%Y-%m-%d %H:%M UTC")
    except (KeyError, ValueError, TypeError):
        return ""


def post_to_row(p: dict, sr: ScoreResult, draft_stub: str = "") -> dict:
    body = (p.get("body_text") or "").strip()
    summary = body[:280] + ("…" if len(body) > 280 else "")
    author = (p.get("author") or "").lstrip("/u/")
    bucket_ru = {
        "relevant": "Релевантно",
        "borderline": "Околорелевантно",
        "irrelevant": "Нерелевантно",
    }[sr.bucket]
    return {
        "Published At": fmt_published(p),
        "Subreddit": f"r/{p.get('subreddit', '')}",
        "Title": (p.get("title") or "").strip(),
        "URL": p.get("url") or "",
        "Summary": summary,
        "Suggested Reply": draft_stub,
        "Final Reply": "",
        "Status": "pending",
        "Score": sr.score,
        "Score Breakdown": format_breakdown(sr.score, sr.breakdown, sr.anti_fits),
        "Bucket": bucket_ru,
        "Manual Override": "",
        "Notes": "",
        "My Hypothesis": "",
        "Author": author,
        "Author Profile URL": f"https://www.reddit.com/u/{author}" if author else "",
        "Upvotes": p.get("upvotes") if p.get("upvotes") is not None else "",
        "Comments": p.get("comments") if p.get("comments") is not None else "",
    }


def make_draft_stub(sr: ScoreResult) -> str:
    """Lightweight stub that gives a future drafter (Claude or human) the
    context needed to write a proper reply. Not the reply itself."""
    parts = [f"DRAFT NEEDED ({sr.bucket})"]
    if sr.use_cases:
        parts.append(f"use_cases: {', '.join(sr.use_cases)}")
    if sr.matched_entities:
        parts.append(f"entities: {', '.join(sr.matched_entities)}")
    return "\n".join(parts)


def style_sheet(ws, freeze_pane: str = "A2") -> None:
    bold = Font(bold=True)
    fill = PatternFill("solid", fgColor="D9E2F3")
    for cell in ws[1]:
        cell.font = bold
        cell.fill = fill
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = freeze_pane


def add_override_validation(ws, col_idx: int) -> None:
    """Add drop-down validation for Manual Override column."""
    dv = DataValidation(
        type="list",
        formula1='"relevant,borderline,irrelevant"',
        allow_blank=True,
    )
    dv.error = "Use one of: relevant, borderline, irrelevant"
    dv.errorTitle = "Invalid Manual Override"
    col_letter = get_column_letter(col_idx)
    dv.add(f"{col_letter}2:{col_letter}10000")
    ws.add_data_validation(dv)


def write_workbook(out_path: Path,
                   draft_rows: list[dict],
                   rejected_rows: list[dict]) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    def add_data_sheet(title: str, rows: list[dict]):
        ws = wb.create_sheet(title=title)
        ws.append(DRAFT_COLS)
        for r in rows:
            ws.append([r.get(c, "") for c in DRAFT_COLS])
        widths = {
            "Published At": 18, "Subreddit": 22, "Title": 38, "URL": 30,
            "Summary": 50, "Suggested Reply": 50, "Final Reply": 50,
            "Status": 10, "Score": 8, "Score Breakdown": 50, "Bucket": 18,
            "Manual Override": 16, "Notes": 28, "My Hypothesis": 40,
            "Author": 22, "Author Profile URL": 30, "Upvotes": 8, "Comments": 8,
        }
        for i, c in enumerate(DRAFT_COLS, 1):
            ws.column_dimensions[get_column_letter(i)].width = widths.get(c, 15)
        style_sheet(ws)
        # Manual Override drop-down
        ovr_col = DRAFT_COLS.index("Manual Override") + 1
        add_override_validation(ws, ovr_col)
        return ws

    add_data_sheet("Drafts", draft_rows)
    add_data_sheet("Auto-rejected", rejected_rows)

    # Гипотезы tab
    ws = wb.create_sheet(title="Гипотезы")
    ws.append(["See voice/rules.md (curated) for the live voice rules + open questions."])
    ws.append([f"Generated at {dt.datetime.now(dt.timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}"])
    ws.column_dimensions["A"].width = 90

    # Reorder: Drafts first, Auto-rejected second, Гипотезы third
    wb._sheets = [wb["Drafts"], wb["Auto-rejected"], wb["Гипотезы"]]
    wb.save(out_path)


def read_workbook(path: Path) -> tuple[list[dict], list[dict]]:
    """Return (drafts_rows, rejected_rows) as list-of-dicts using DRAFT_COLS."""
    wb = load_workbook(path)
    out: dict[str, list[dict]] = {"Drafts": [], "Auto-rejected": []}
    for sheet_name in out:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        header = list(rows[0])
        for r in rows[1:]:
            if r is None or all(c is None or c == "" for c in r):
                continue
            out[sheet_name].append({h: (r[i] if i < len(r) else "") for i, h in enumerate(header)})
    return out["Drafts"], out["Auto-rejected"]


# =============================================================================
# Commands
# =============================================================================

def cmd_scan(args) -> int:
    scoring = load_yaml(SCORING_PATH)
    entities = load_yaml(ENTITIES_PATH)
    vendor_authors = set(entities.get("vendor_blogs", []) or [])
    scorer = Scorer(scoring, entities, vendor_authors)

    print(f"=== scan: last {args.hours}h ===", file=sys.stderr)
    if not SUBS:
        print("SUBS is empty — edit orchestrator.py to add subreddits to scan.",
              file=sys.stderr)
        return 1
    all_posts: list[dict] = []
    for slug, pages in SUBS:
        print(f"[{slug}] (pages={pages})...", file=sys.stderr)
        posts = pull_sub(slug, pages)
        recent = filter_recent(posts, args.hours)
        for p in recent:
            p["_score"] = scorer.score_post(p)
        all_posts.extend(recent)
        print(f"    pulled={len(posts)} recent={len(recent)}", file=sys.stderr)

    mark_cross_posts(all_posts, scorer)

    # Bucket
    buckets = {"relevant": [], "borderline": [], "irrelevant": []}
    for p in all_posts:
        buckets[p["_score"].bucket].append(p)

    print(f"\nbuckets: relevant={len(buckets['relevant'])}  "
          f"borderline={len(buckets['borderline'])}  "
          f"irrelevant={len(buckets['irrelevant'])}", file=sys.stderr)

    # Enrich relevant + borderline only (irrelevant stays cheap)
    to_enrich = buckets["relevant"] + buckets["borderline"]
    if to_enrich and not args.skip_enrich:
        print(f"\nenriching {len(to_enrich)} permalinks...", file=sys.stderr)
        enriched = enrich_posts(to_enrich)
        for p in to_enrich:
            e = enriched.get(p["url"], {})
            p["upvotes"] = e.get("upvotes")
            p["comments"] = e.get("comments")

    # Build rows
    draft_rows = []
    for bucket_key in ("relevant", "borderline"):
        for p in sorted(buckets[bucket_key], key=lambda q: -q["_score"].score):
            stub = make_draft_stub(p["_score"])
            draft_rows.append(post_to_row(p, p["_score"], draft_stub=stub))

    rejected_rows = []
    for p in sorted(buckets["irrelevant"], key=lambda q: q.get("subreddit") or ""):
        rejected_rows.append(post_to_row(p, p["_score"], draft_stub=""))

    out = HERE / args.out
    write_workbook(out, draft_rows, rejected_rows)
    print(f"\n→ {out} ({out.stat().st_size:,} bytes, "
          f"{len(draft_rows)} drafts, {len(rejected_rows)} rejected)", file=sys.stderr)
    return 0


def _bucket_to_key(label: str) -> str | None:
    return {
        "Релевантно": "relevant",
        "Околорелевантно": "borderline",
        "Нерелевантно": "irrelevant",
    }.get(label)


def cmd_sync_overrides(args) -> int:
    """Read xlsx, apply Manual Override, log classification_feedback,
    move rows between sheets, re-write xlsx."""
    path = HERE / args.workbook
    if not path.exists():
        print(f"workbook not found: {path}", file=sys.stderr)
        return 1

    drafts, rejected = read_workbook(path)
    print(f"loaded: {len(drafts)} drafts, {len(rejected)} rejected", file=sys.stderr)

    # Find rows with non-empty Manual Override
    moved = 0
    new_drafts = []
    new_rejected = []
    for source_name, rows, target_label in [
        ("Drafts", drafts, "drafts"),
        ("Auto-rejected", rejected, "auto-rejected"),
    ]:
        for row in rows:
            ovr = (row.get("Manual Override") or "").strip().lower()
            if ovr in {"relevant", "borderline", "irrelevant"}:
                old_bucket = _bucket_to_key(row.get("Bucket") or "")
                if old_bucket != ovr:
                    moved += 1
                    append_lesson({
                        "type": "classification_feedback",
                        "ts": dt.datetime.now(dt.timezone.utc).isoformat(),
                        "url": row.get("URL"),
                        "subreddit": row.get("Subreddit"),
                        "title": row.get("Title"),
                        "body_text_excerpt": (row.get("Summary") or "")[:300],
                        "my_bucket": old_bucket,
                        "my_score": row.get("Score"),
                        "my_score_breakdown": row.get("Score Breakdown"),
                        "my_hypothesis": row.get("My Hypothesis"),
                        "user_override": ovr,
                        "source_sheet": source_name,
                    })
                # Apply: change Bucket cell, clear Manual Override (consumed)
                row["Bucket"] = {
                    "relevant": "Релевантно",
                    "borderline": "Околорелевантно",
                    "irrelevant": "Нерелевантно",
                }[ovr]
                row["Manual Override"] = ""

            # Route by NEW bucket
            if (row.get("Bucket") or "") == "Нерелевантно":
                new_rejected.append(row)
            else:
                new_drafts.append(row)

    # Re-write
    write_workbook(path, new_drafts, new_rejected)
    print(f"\nmoved {moved} posts; lessons appended → {LESSONS_PATH}", file=sys.stderr)
    print(f"→ {path}", file=sys.stderr)
    return 0


def _diff_summary(suggested: str, final: str) -> dict:
    s_lines = (suggested or "").splitlines()
    f_lines = (final or "").splitlines()
    s_sentences = re.split(r"(?<=[.!?])\s+", (suggested or "").strip())
    f_sentences = re.split(r"(?<=[.!?])\s+", (final or "").strip())
    s_phrases = set(_phrases(suggested))
    f_phrases = set(_phrases(final))
    return {
        "length_change": f"{len(s_sentences)} → {len(f_sentences)} sentences "
                         f"({len(suggested)} → {len(final)} chars)",
        "removed_phrases": sorted(s_phrases - f_phrases)[:10],
        "added_phrases": sorted(f_phrases - s_phrases)[:10],
        "unified_diff": "\n".join(difflib.unified_diff(
            s_lines, f_lines, lineterm="", n=1))[:2000],
    }


def _phrases(text: str) -> list[str]:
    if not text:
        return []
    sents = re.split(r"(?<=[.!?])\s+", text.strip())
    out = []
    for s in sents:
        s = s.strip().strip('"\'').strip()
        if 4 < len(s) < 200:
            out.append(s)
    return out


def cmd_sync_voice(args) -> int:
    """Read xlsx, diff Suggested↔Final, append voice_feedback to lessons.jsonl,
    refresh examples.jsonl with the latest 10 actual replies."""
    path = HERE / args.workbook
    if not path.exists():
        print(f"workbook not found: {path}", file=sys.stderr)
        return 1

    drafts, _ = read_workbook(path)
    new_examples = []
    new_lessons = 0
    for row in drafts:
        suggested = (row.get("Suggested Reply") or "").strip()
        final = (row.get("Final Reply") or "").strip()
        if not final:
            continue
        if final == suggested:
            continue
        # Skip drafts that are still placeholder stubs ("DRAFT NEEDED ...")
        if suggested.startswith("DRAFT NEEDED"):
            suggested = ""

        new_lessons += 1
        diff = _diff_summary(suggested, final)
        append_lesson({
            "type": "voice_feedback",
            "ts": dt.datetime.now(dt.timezone.utc).isoformat(),
            "url": row.get("URL"),
            "subreddit": row.get("Subreddit"),
            "title": row.get("Title"),
            "category": {
                "bucket": _bucket_to_key(row.get("Bucket") or ""),
                "score": row.get("Score"),
                "score_breakdown": row.get("Score Breakdown"),
            },
            "my_draft": suggested,
            "user_final": final,
            "diff_summary": diff,
        })
        new_examples.append({
            "url": row.get("URL"),
            "subreddit": row.get("Subreddit"),
            "context": row.get("Score Breakdown"),
            "final_reply": final,
        })

    # Refresh examples.jsonl: keep last 10 (newest first)
    existing = []
    if EXAMPLES_PATH.exists():
        existing = [json.loads(l) for l in EXAMPLES_PATH.read_text(encoding="utf-8").splitlines() if l.strip()]
    merged = new_examples + existing
    # De-dup by URL
    seen = set()
    final_list = []
    for ex in merged:
        if ex["url"] in seen:
            continue
        seen.add(ex["url"])
        final_list.append(ex)
        if len(final_list) >= 10:
            break

    EXAMPLES_PATH.write_text(
        "\n".join(json.dumps(e, ensure_ascii=False) for e in final_list) + ("\n" if final_list else ""),
        encoding="utf-8",
    )

    print(f"voice_feedback added: {new_lessons}", file=sys.stderr)
    print(f"examples.jsonl: {len(final_list)} entries", file=sys.stderr)
    return 0


# =============================================================================
# CLI
# =============================================================================

def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__.split("\n", 1)[0])
    sub = ap.add_subparsers(dest="cmd", required=True)

    sp = sub.add_parser("scan", help="Pull RSS, score, bucket, enrich, write xlsx")
    sp.add_argument("--hours", type=int, default=72)
    sp.add_argument("--out", default="scan.xlsx")
    sp.add_argument("--skip-enrich", action="store_true",
                    help="Skip SB-CDP enrich step (faster smoke test)")
    sp.set_defaults(func=cmd_scan)

    sp = sub.add_parser("sync-overrides",
                        help="Read xlsx, apply Manual Override, log lessons")
    sp.add_argument("--workbook", default="scan.xlsx")
    sp.set_defaults(func=cmd_sync_overrides)

    sp = sub.add_parser("sync-voice",
                        help="Read xlsx, diff Suggested↔Final, log lessons + examples")
    sp.add_argument("--workbook", default="scan.xlsx")
    sp.set_defaults(func=cmd_sync_voice)

    args = ap.parse_args()
    return args.func(args)


if __name__ == "__main__":
    sys.exit(main())
