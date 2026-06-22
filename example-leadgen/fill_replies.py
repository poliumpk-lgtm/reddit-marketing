"""Fill the Suggested Reply column in scan.xlsx with hand-written drafts.

Workflow: after `python orchestrator.py scan` produces scan.xlsx, populate
the DRAFTS dict below with `{post_url: reply_text}` and run this script.
It writes each reply into the matching row's Suggested Reply cell.

Conventions for drafts (see voice/rules.md):
  - 3-5 sentences, peer tone, technical, ends with a counter-question.
  - For cross-posters whose score got `cross_post_duplicate`, use a draft
    starting with "SKIP — " so the row gets marked but no reply is sent.
"""
import sys
from openpyxl import load_workbook
from openpyxl.styles import Alignment

WB_PATH = "scan.xlsx"

# URL → suggested reply text. Add entries after each scan.
# Example:
#   "https://www.reddit.com/r/YourSub/comments/abc123/title/":
#       "Your 3-5 sentence reply ending with a counter-question?",
DRAFTS: dict[str, str] = {}


def main() -> int:
    wb = load_workbook(WB_PATH)
    ws = wb["Drafts"]
    header = [c.value for c in ws[1]]
    url_idx = header.index("URL")
    sr_idx = header.index("Suggested Reply")

    filled = 0
    skipped = 0
    not_found = 0
    for row in ws.iter_rows(min_row=2):
        url = row[url_idx].value
        if not url:
            continue
        draft = DRAFTS.get(url)
        if draft is None:
            not_found += 1
            print(f"  ! no draft for {url}", file=sys.stderr)
            continue
        if draft.startswith("SKIP"):
            skipped += 1
        else:
            filled += 1
        cell = row[sr_idx]
        cell.value = draft
        cell.alignment = Alignment(vertical="top", wrap_text=True)

    wb.save(WB_PATH)
    print(f"\nfilled: {filled}, skipped (cross-post): {skipped}, no_draft: {not_found}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
